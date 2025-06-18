from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from .models import Kategori, Eslesme, Bildirim
from .forms import KategoriForm, EslesmeForm. CustomUserCreationForm
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
from django.http import FileResponse
from ratelimit.decorators import ratelimit

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

CHROMEDRIVER_PATH = r"C:\Users\ONUR\Desktop\Yeni klasör (4)\chromedriver.exe"

def export_to_excel_dark(df, excel_path):
    columns = list(df.columns)
    akakce_link_kolonu = "Akakçe Linki"
    site_link_kolonu = "Sitenizdeki Link"
    new_columns = [c for c in columns if c not in [akakce_link_kolonu, site_link_kolonu]] + [akakce_link_kolonu, site_link_kolonu]
    df = df.reindex(columns=new_columns)
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="FF23235A", end_color="FF23235A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", name="Segoe UI", size=12)
    zebra1 = PatternFill(start_color="FF191940", end_color="FF191940", fill_type="solid")
    zebra2 = PatternFill(start_color="FF23235A", end_color="FF23235A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = zebra1 if i % 2 == 0 else zebra2
        for cell in row:
            cell.fill = fill
            cell.font = Font(name="Segoe UI", size=11, color="FFEEEEFF")
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
                if val is None: val = ""
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        new_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = new_width

    for idx, cell in enumerate(ws[1], 1):
        if cell.value in [akakce_link_kolonu, site_link_kolonu]:
            ws.column_dimensions[cell.column_letter].hidden = True

    wb.save(excel_path)

def home_view(request):
    return render(request, "home.html")

def register(request):
    if request.user.is_authenticated:
        return redirect('profile')
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Kayıt başarılı! Giriş yapabilirsiniz.")
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, "register.html", {"form": form})

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect("fiyat_cek")
        else:
            messages.error(request, "Kullanıcı adı veya şifre hatalı.")
    return render(request, "login.html")

def logout_view(request):
    logout(request)
    return redirect("login")

@login_required
def profile(request):
    kategoriler = Kategori.objects.filter(user=request.user)
    kategori_id = request.GET.get("kategori")
    aktif_kategori = None
    eslesmeler = []
    eslesme_form = EslesmeForm()
    kategori_form = KategoriForm()

    if kategori_id:
        aktif_kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
        eslesmeler = Eslesme.objects.filter(kategori=aktif_kategori)

    if request.method == "POST":
        if "ekle_kategori" in request.POST:
            kategori_form = KategoriForm(request.POST)
            if kategori_form.is_valid():
                kategori = kategori_form.save(commit=False)
                kategori.user = request.user
                kategori.save()
                bildirim_ekle(request.user, f"Yeni kategori eklendi: {kategori.isim}")
                messages.success(request, "Kategori eklendi.")
                return redirect(f"{request.path}?kategori={kategori.id}")
        elif "sil_kategori" in request.POST:
            sil_id = request.POST.get("sil_kategori_id")
            sil_kat = Kategori.objects.filter(id=sil_id, user=request.user)
            if sil_kat.exists():
                kategori_ismi = sil_kat[0].isim
                sil_kat.delete()
                bildirim_ekle(request.user, f"Kategori silindi: {kategori_ismi}")
                messages.success(request, "Kategori silindi.")
                kalan = Kategori.objects.filter(user=request.user).first()
                return redirect(f"{request.path}?kategori={kalan.id if kalan else ''}")
        elif "eslesme_ekle" in request.POST and kategori_id:
            eslesme_form = EslesmeForm(request.POST)
            if eslesme_form.is_valid():
                eslesme = eslesme_form.save(commit=False)
                eslesme.kategori = aktif_kategori
                eslesme.save()
                bildirim_ekle(request.user, f"Yeni eşleşme eklendi: {eslesme.akakce_link}")
                messages.success(request, "Eşleşme eklendi.")
                return redirect(f"{request.path}?kategori={aktif_kategori.id}")
        elif "eslesme_sil" in request.POST:
            sil_id = request.POST.get("eslesme_sil_id")
            eslesme = Eslesme.objects.filter(id=sil_id, kategori__user=request.user)
            if eslesme.exists():
                kategori_id = eslesme[0].kategori.id
                bildirim_ekle(request.user, f"Eşleşme silindi: {eslesme[0].akakce_link}")
                eslesme.delete()
                messages.success(request, "Eşleşme silindi.")
                return redirect(f"{request.path}?kategori={kategori_id}")
        elif "fiyatlari_cek" in request.POST and kategori_id:
            kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
            eslesmeler = Eslesme.objects.filter(kategori=kategori)
            chrome_options = Options()
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)")
            chrome_options.add_argument("--headless")
            service = Service(CHROMEDRIVER_PATH)
            driver = webdriver.Chrome(service=service, options=chrome_options)
            data = []
            for eslesme in eslesmeler:
                akakce_url = eslesme.akakce_link
                site_url = eslesme.site_link
                urun_adi = ""
                akakce_fiyat = ""
                akakce_magaza = ""
                site_fiyat = ""
                fiyat_farki = ""
                try:
                    driver.get(akakce_url)
                    driver.implicitly_wait(8)
                    urun_adi = driver.find_element(By.TAG_NAME, "h1").text.strip()
                    fiyat = driver.find_element(By.CSS_SELECTOR, "ul#PL li a.iC.xt_v8 .pt_v8").text.strip()
                    magaza = driver.find_element(By.CSS_SELECTOR, "ul#PL li a.iC.xt_v8 .v_v8").text.strip()
                    akakce_fiyat = fiyat.replace("\n", "").replace("\r", "")
                    akakce_magaza = magaza
                except Exception:
                    urun_adi = "Ürün adı bulunamadı"
                    akakce_fiyat = "Fiyat bulunamadı"
                    akakce_magaza = "Mağaza bulunamadı"
                if site_url:
                    try:
                        driver.get(site_url)
                        driver.implicitly_wait(6)
                        site_fiyat = driver.find_element(By.CSS_SELECTOR, "div.product-price-new").text.strip()
                    except Exception:
                        site_fiyat = "Fiyat bulunamadı"
                try:
                    af = float(str(akakce_fiyat).replace(".", "").replace(",", ".").split()[0])
                    sf = float(str(site_fiyat).replace(".", "").replace(",", ".").split()[0])
                    fiyat_farki = round(sf - af, 2)
                except Exception:
                    fiyat_farki = ""
                data.append({
                    "Ürün Adı": urun_adi,
                    "Akakçe Fiyatı": akakce_fiyat,
                    "Akakçe Mağaza": akakce_magaza,
                    "Akakçe Linki": akakce_url,
                    "Sitenizdeki Link": site_url,
                    "Site Fiyatınız": site_fiyat,
                    "Fiyat Farkı": fiyat_farki,
                })
            driver.quit()
            df = pd.DataFrame(data)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                export_to_excel_dark(df, tmp.name)
                response = FileResponse(open(tmp.name, "rb"), as_attachment=True, filename=f"fiyatlar_{kategori.isim}.xlsx")
                bildirim_ekle(request.user, f"{kategori.isim} kategorisi için fiyatlar Excel olarak indirildi.")
                return response
        return redirect(f"{request.path}?kategori={kategori_id or ''}")

    context = {
        "kategoriler": kategoriler,
        "aktif_kategori": aktif_kategori,
        "eslesmeler": eslesmeler,
        "eslesme_form": eslesme_form,
        "kategori_form": kategori_form,
        "user": request.user,
        "messages": messages.get_messages(request),
    }
    return render(request, "profile.html", context)

@ratelimit(key='user', rate='10/m', block=True)
@login_required
def fiyat_cek(request):
    kategoriler = Kategori.objects.filter(user=request.user)
    kategori_id = request.GET.get("kategori")
    aktif_kategori = None
    eslesmeler = []
    eslesme_form = EslesmeForm()
    kategori_form = KategoriForm()

    if kategori_id:
        aktif_kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
        eslesmeler = Eslesme.objects.filter(kategori=aktif_kategori)

    if request.method == "POST":
        if "ekle_kategori" in request.POST:
            kategori_form = KategoriForm(request.POST)
            if kategori_form.is_valid():
                kategori = kategori_form.save(commit=False)
                kategori.user = request.user
                kategori.save()
                bildirim_ekle(request.user, f"Yeni kategori eklendi: {kategori.isim}")
                messages.success(request, "Kategori eklendi.")
                return redirect(f"{request.path}?kategori={kategori.id}")
        elif "sil_kategori" in request.POST:
            sil_id = request.POST.get("sil_kategori_id")
            sil_kat = Kategori.objects.filter(id=sil_id, user=request.user)
            if sil_kat.exists():
                kategori_ismi = sil_kat[0].isim
                sil_kat.delete()
                bildirim_ekle(request.user, f"Kategori silindi: {kategori_ismi}")
                messages.success(request, "Kategori silindi.")
                kalan = Kategori.objects.filter(user=request.user).first()
                return redirect(f"{request.path}?kategori={kalan.id if kalan else ''}")
        elif "eslesme_ekle" in request.POST and kategori_id:
            eslesme_form = EslesmeForm(request.POST)
            if eslesme_form.is_valid():
                eslesme = eslesme_form.save(commit=False)
                eslesme.kategori = aktif_kategori
                eslesme.save()
                bildirim_ekle(request.user, f"Yeni eşleşme eklendi: {eslesme.akakce_link}")
                messages.success(request, "Eşleşme eklendi.")
                return redirect(f"{request.path}?kategori={aktif_kategori.id}")
        elif "eslesme_sil" in request.POST:
            sil_id = request.POST.get("eslesme_sil_id")
            eslesme = Eslesme.objects.filter(id=sil_id, kategori__user=request.user)
            if eslesme.exists():
                kategori_id = eslesme[0].kategori.id
                bildirim_ekle(request.user, f"Eşleşme silindi: {eslesme[0].akakce_link}")
                eslesme.delete()
                messages.success(request, "Eşleşme silindi.")
                return redirect(f"{request.path}?kategori={kategori_id}")
        elif "fiyatlari_cek" in request.POST and kategori_id:
            kategori = get_object_or_404(Kategori, id=kategori_id, user=request.user)
            eslesmeler = Eslesme.objects.filter(kategori=kategori)
            chrome_options = Options()
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64)")
            chrome_options.add_argument("--headless")
            service = Service(CHROMEDRIVER_PATH)
            driver = webdriver.Chrome(service=service, options=chrome_options)
            data = []
            for eslesme in eslesmeler:
                akakce_url = eslesme.akakce_link
                site_url = eslesme.site_link
                urun_adi = ""
                akakce_fiyat = ""
                akakce_magaza = ""
                site_fiyat = ""
                fiyat_farki = ""
                try:
                    driver.get(akakce_url)
                    driver.implicitly_wait(8)
                    urun_adi = driver.find_element(By.TAG_NAME, "h1").text.strip()
                    fiyat = driver.find_element(By.CSS_SELECTOR, "ul#PL li a.iC.xt_v8 .pt_v8").text.strip()
                    magaza = driver.find_element(By.CSS_SELECTOR, "ul#PL li a.iC.xt_v8 .v_v8").text.strip()
                    akakce_fiyat = fiyat.replace("\n", "").replace("\r", "")
                    akakce_magaza = magaza
                except Exception:
                    urun_adi = "Ürün adı bulunamadı"
                    akakce_fiyat = "Fiyat bulunamadı"
                    akakce_magaza = "Mağaza bulunamadı"
                if site_url:
                    try:
                        driver.get(site_url)
                        driver.implicitly_wait(6)
                        site_fiyat = driver.find_element(By.CSS_SELECTOR, "div.product-price-new").text.strip()
                    except Exception:
                        site_fiyat = "Fiyat bulunamadı"
                try:
                    af = float(str(akakce_fiyat).replace(".", "").replace(",", ".").split()[0])
                    sf = float(str(site_fiyat).replace(".", "").replace(",", ".").split()[0])
                    fiyat_farki = round(sf - af, 2)
                except Exception:
                    fiyat_farki = ""
                data.append({
                    "Ürün Adı": urun_adi,
                    "Akakçe Fiyatı": akakce_fiyat,
                    "Akakçe Mağaza": akakce_magaza,
                    "Akakçe Linki": akakce_url,
                    "Sitenizdeki Link": site_url,
                    "Site Fiyatınız": site_fiyat,
                    "Fiyat Farkı": fiyat_farki,
                })
            driver.quit()
            df = pd.DataFrame(data)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                export_to_excel_dark(df, tmp.name)
                response = FileResponse(open(tmp.name, "rb"), as_attachment=True, filename=f"fiyatlar_{kategori.isim}.xlsx")
                bildirim_ekle(request.user, f"{kategori.isim} kategorisi için fiyatlar Excel olarak indirildi.")
                return response
        return redirect(f"{request.path}?kategori={kategori_id or ''}")

    context = {
        "kategoriler": kategoriler,
        "aktif_kategori": aktif_kategori,
        "eslesmeler": eslesmeler,
        "eslesme_form": eslesme_form,
        "kategori_form": kategori_form,
        "user": request.user,
        "messages": messages.get_messages(request),
    }
    return render(request, "fiyat_cek.html", context)

@login_required
def settings_view(request):
    kategoriler = Kategori.objects.filter(user=request.user)
    default_category_id = getattr(request.user.userprofile, "default_category_id", None)
    theme = request.session.get("theme", "light")

    if request.method == "POST":
        email = request.POST.get("email")
        if email and email != request.user.email:
            request.user.email = email
            request.user.save()
            messages.success(request, "E-posta güncellendi.")
        default_category_id = request.POST.get("default_category")
        if default_category_id and default_category_id.isdigit():
            request.user.userprofile.default_category_id = int(default_category_id)
            request.user.userprofile.save()
            messages.success(request, "Varsayılan kategori güncellendi.")
        theme = request.POST.get("theme", "light")
        request.session["theme"] = theme
        messages.success(request, "Tema tercihi kaydedildi.")

    context = {
        "user": request.user,
        "kategoriler": kategoriler,
        "default_category_id": default_category_id,
        "theme": theme,
    }
    return render(request, "ayarlar.html", context)

def bildirim_ekle(user, mesaj):
    Bildirim.objects.create(user=user, mesaj=mesaj)

@login_required
def bildirim_merkezi(request):
    bildirimler = request.user.bildirimler.order_by('-olusturma_zamani')
    if request.method == "POST":
        request.user.bildirimler.filter(okundu=False).update(okundu=True)
    return render(request, "bildirimler.html", {"bildirimler": bildirimler})